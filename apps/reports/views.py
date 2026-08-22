"""Operator/Admin dashboard — overview, requests, quotes, status, assign, pickup, reports/export."""

import calendar
import contextlib
import csv
from datetime import timedelta
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count, DecimalField, Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.accounts.models import User
from apps.accounts.views import role_required, staff_required
from apps.branches.models import Branch
from apps.intake.models import IntakeRequest
from apps.notifications.models import EmailLog
from apps.notifications.services import (
    notify_pickup_scheduled,
    notify_quote_sent,
    notify_status_changed,
)
from apps.quotes.forms import AssignForm, PickupForm, QuotationForm, StatusChangeForm
from apps.quotes.models import Pickup, StatusHistory

from .filters import IntakeRequestFilter

EXCEL_COLUMNS = [
    "Код",
    "Огноо",
    "Хэрэглэгч",
    "Утас",
    "Email",
    "Компани",
    "Салбар",
    "Статус",
    "Хариуцагч",
    "Төхөөрөмжийн тоо",
    "Хүссэн үнэ",
    "Бодит үнэ",
]


@staff_required
def overview(request):
    today = timezone.localdate()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    total = IntakeRequest.objects.count()
    today_count = IntakeRequest.objects.filter(created_at__date=today).count()
    week_count = IntakeRequest.objects.filter(created_at__date__gte=week_ago).count()
    week_quoted = (
        IntakeRequest.objects.filter(created_at__date__gte=week_ago, quotes__isnull=False)
        .distinct()
        .count()
    )
    purchased_month = IntakeRequest.objects.filter(
        status=IntakeRequest.Status.PURCHASED, updated_at__date__gte=month_ago
    )
    purchased_count = purchased_month.count()
    purchased_amount = Pickup.objects.filter(intake_request__in=purchased_month).aggregate(
        total=Coalesce(Sum("actual_buy_price"), 0, output_field=DecimalField())
    )["total"]

    status_labels = dict(IntakeRequest.Status.choices)
    by_status = [
        {
            "status": row["status"],
            "label": status_labels.get(row["status"], row["status"]),
            "c": row["c"],
        }
        for row in IntakeRequest.objects.values("status").annotate(c=Count("id")).order_by("-c")
    ]
    by_branch = list(
        IntakeRequest.objects.exclude(preferred_branch=None)
        .values("preferred_branch__name")
        .annotate(c=Count("id"))
        .order_by("-c")
    )
    by_category = list(
        IntakeRequest.objects.exclude(items__category__name=None)
        .values("items__category__name")
        .annotate(c=Count("id"))
        .order_by("-c")
    )

    recent = IntakeRequest.objects.select_related("preferred_branch", "assigned_to").order_by(
        "-created_at"
    )[:10]

    return render(
        request,
        "dashboard/overview.html",
        {
            "total": total,
            "today_count": today_count,
            "week_count": week_count,
            "week_quoted": week_quoted,
            "purchased_count": purchased_count,
            "purchased_amount": purchased_amount,
            "by_status": by_status,
            "by_branch": by_branch,
            "by_category": by_category,
            "recent": recent,
        },
    )


@staff_required
def delivery_map(request):
    """Хүргэлт — дуусаагүй, хүргэлтээр авах хүсэлтүүдийг газрын зураг дээр харуулна."""
    open_requests = (
        IntakeRequest.objects.filter(
            pickup_required=True, status__in=IntakeRequest.OPEN_STATUSES
        )
        .select_related("preferred_branch", "assigned_to")
        .order_by("-created_at")
    )
    points = [
        {
            "code": r.request_code,
            "lat": float(r.pickup_lat),
            "lng": float(r.pickup_lng),
            "name": r.contact_name,
            "phone": r.contact_phone,
            "address": ", ".join(p for p in [r.city, r.district, r.address_line] if p),
            "status": r.get_status_display(),
            "url": r.get_absolute_url(),
        }
        for r in open_requests
        if r.has_location
    ]
    return render(
        request,
        "dashboard/delivery.html",
        {
            "requests": open_requests,
            "points": points,
            "mapped_count": len(points),
        },
    )


PERIOD_CHOICES = [
    ("this_month", "Энэ сар"),
    ("last_month", "Өмнөх сар"),
    ("last_3m", "Сүүлийн 3 сар"),
    ("last_6m", "Сүүлийн 6 сар"),
]


def _month_first(d):
    return d.replace(day=1)


def _month_last(d):
    return d.replace(day=calendar.monthrange(d.year, d.month)[1])


def _shift_months(d, n):
    """First day of the month `n` months before `d` (n >= 0)."""
    month = d.month - n
    year = d.year
    while month <= 0:
        month += 12
        year -= 1
    return d.replace(year=year, month=month, day=1)


def _period_range(period, today):
    """Return (start_date, end_date) for a named preset period."""
    if period == "this_month":
        return _month_first(today), _month_last(today)
    if period == "last_month":
        prev = _shift_months(today, 1)
        return prev, _month_last(prev)
    if period == "last_3m":
        return _shift_months(today, 2), _month_last(today)
    if period == "last_6m":
        return _shift_months(today, 5), _month_last(today)
    return None, None


# Жагсаалтад нэг хуудсанд харуулах мөрийн тоо. "all" = хуудаслалтгүй, бүгд.
PER_PAGE_CHOICES = [("10", "10"), ("25", "25"), ("50", "50"), ("all", "Бүгд")]
DEFAULT_PER_PAGE = "25"


@staff_required
def request_list(request):
    params = request.GET.copy()
    period = params.get("period", "")
    start, end = _period_range(period, timezone.localdate())
    if start and end:
        params["created_at__gte"] = start.isoformat()
        params["created_at__lte"] = end.isoformat()

    qs = IntakeRequest.objects.select_related("preferred_branch", "assigned_to").order_by(
        "-created_at"
    )
    f = IntakeRequestFilter(params, queryset=qs)

    per_page = params.get("per_page") or DEFAULT_PER_PAGE
    if per_page not in dict(PER_PAGE_CHOICES):
        per_page = DEFAULT_PER_PAGE

    total = f.qs.count()
    if per_page == "all":
        page_obj = None
        rows = f.qs
    else:
        page_obj = Paginator(f.qs, int(per_page)).get_page(params.get("page"))
        rows = page_obj.object_list

    return render(
        request,
        "dashboard/request_list.html",
        {
            "filter": f,
            "requests": rows,
            "page_obj": page_obj,
            "per_page": per_page,
            "per_page_choices": PER_PAGE_CHOICES,
            "total": total,
            "period": period,
            "period_choices": PERIOD_CHOICES,
        },
    )


@staff_required
def request_detail(request, code):
    intake = get_object_or_404(
        IntakeRequest.objects.select_related("preferred_branch", "assigned_to", "submitted_by"),
        request_code=code,
    )
    latest_quote = _current_quote(intake)
    items = list(intake.items.prefetch_related("images").all())
    return render(
        request,
        "dashboard/request_detail.html",
        {
            "request_obj": intake,
            "items": items,
            # Lightbox-д зориулж бүх зургийг нэг эрэмбэтэй жагсаалтаар — жижиг
            # зураг дээр дарахад тэндээсээ өмнөх/дараах руу гүйлгэнэ.
            "gallery_images": [img.image.url for item in items for img in item.images.all()],
            "history": intake.history.all(),
            "latest_quote": latest_quote,
            "pickup": getattr(intake, "pickup", None),
            # Мэдэгдэл явсан эсэх / алдааг оператор шууд харна.
            "email_logs": intake.email_logs.all()[:5],
            # Одоогийн саналыг форм дээр урьдчилан дүүргэнэ — засаад дахин илгээнэ.
            # Ижил загварын өмнөх үнэ — оператор шинэ үнээ түүнд тааруулна.
            "similar_quotes": _similar_priced_requests(intake),
            "quote_form": QuotationForm(instance=latest_quote),
            "status_form": StatusChangeForm(initial={"new_status": intake.status}),
            "assign_form": AssignForm(initial={"assigned_to": intake.assigned_to}),
        },
    )


@staff_required
def add_quote(request, code):
    intake = get_object_or_404(IntakeRequest, request_code=code)
    if request.method != "POST":
        return redirect("dashboard:request_detail", code=code)
    # Хүсэлт тутамд ганц үнэ санал байна — дахин илгээвэл хуучныг нь шинэчилнэ.
    form = QuotationForm(request.POST, instance=_current_quote(intake))
    if form.is_valid():
        q = form.save(commit=False)
        q.intake_request = intake
        q.quoted_by = request.user
        q.sent_to_customer_at = timezone.now()
        q.save()
        notify_quote_sent(q)
        if intake.status == IntakeRequest.Status.NEW:
            _change_status(intake, IntakeRequest.Status.PRICE_SENT, request.user, "Үнэ санал илгээв")
        if intake.contact_email:
            # Имэйл нь background-д илгээгддэг тул энд амжилтыг батлах боломжгүй —
            # үр дүнг хажуугийн "Илгээсэн имэйл" самбараас харна.
            messages.success(
                request,
                f"Үнэ санал хадгалагдлаа. {intake.contact_email} хаяг руу имэйл "
                "илгээж байна — үр дүнг хуудасны баруун талаас шалгана уу.",
            )
        else:
            messages.warning(
                request,
                "Үнэ санал хадгалагдлаа. Хэрэглэгч имэйл хаяг үлдээгээгүй тул захиа явуулаагүй.",
            )
    else:
        messages.error(request, "Үнэ санал хадгалахад алдаа гарлаа.")
    return redirect("dashboard:request_detail", code=code)


def _current_quote(intake):
    """Хүсэлтийн одоо мөрдөгдөж буй үнэ санал (байхгүй бол None)."""
    return intake.quotes.order_by("-created_at").first()


# Дэлгэрэнгүй хуудсанд лавлагаа болгож харуулах өмнөх үнэ саналын мөрийн тоо.
SIMILAR_QUOTE_LIMIT = 10


def _similar_priced_requests(intake, limit=SIMILAR_QUOTE_LIMIT):
    """Ижил бренд + ангилалтай, аль хэдийн үнэ өгөгдсөн бусад хүсэлтүүд.

    Загвар яг таарахыг шаардвал (iPhone 13 ≠ iPhone 13 Pro) жагсаалт бараг
    үргэлж хоосон гардаг байсан. Тиймээс "Apple гар утас" гэсэн түвшинд —
    бренд, ангилал хоёр нь тухайн хүсэлттэй адил бол — тааруулж, хамгийн сүүлд
    ирсэн хүсэлтээс нь эхлүүлж жагсаана. Загвар, төлөв нь мөрөндөө харагдах тул
    оператор аль нь яг таарч байгааг, ямар үнээр хэлцэл болсныг өөрөө жиших
    боломжтой. Мөр бүр нь тухайн хүсэлтийн дэлгэрэнгүй рүү холбогдоно.
    """
    wanted = {
        (item.brand.strip().lower(), item.category_id)
        for item in intake.items.all()
        if item.brand.strip()
    }
    if not wanted:
        return []

    match = Q()
    for brand, category_id in wanted:
        match |= Q(items__brand__iexact=brand, items__category_id=category_id)

    others = (
        IntakeRequest.objects.filter(match)
        .exclude(pk=intake.pk)
        .filter(quotes__isnull=False)
        .distinct()
        .prefetch_related("items", "quotes")
        .order_by("-created_at")[:limit]
    )

    rows = []
    for other in others:
        quote = max(other.quotes.all(), key=lambda q: q.created_at, default=None)
        if quote is None:
            continue
        device = next(
            (
                it
                for it in other.items.all()
                if (it.brand.strip().lower(), it.category_id) in wanted
            ),
            None,
        )
        rows.append({"request": other, "quote": quote, "device": device})
    return rows


@staff_required
def change_status(request, code):
    intake = get_object_or_404(IntakeRequest, request_code=code)
    if request.method != "POST":
        return redirect("dashboard:request_detail", code=code)
    form = StatusChangeForm(request.POST)
    if form.is_valid():
        new_status = form.cleaned_data["new_status"]
        comment = form.cleaned_data["comment"]
        if new_status != intake.status:
            _change_status(intake, new_status, request.user, comment)
            messages.success(request, "Статус шинэчлэгдлээ.")
        else:
            messages.info(request, "Статус өөрчлөгдөөгүй.")
    return redirect("dashboard:request_detail", code=code)


def _change_status(intake, new_status, by_user, comment):
    old = intake.status
    intake.status = new_status
    intake.save(update_fields=["status", "updated_at"])
    StatusHistory.objects.create(
        intake_request=intake,
        old_status=old,
        new_status=new_status,
        comment=comment,
        changed_by=by_user,
    )
    notify_status_changed(intake, old, new_status, comment)


@staff_required
def assign(request, code):
    intake = get_object_or_404(IntakeRequest, request_code=code)
    if request.method != "POST":
        return redirect("dashboard:request_detail", code=code)
    form = AssignForm(request.POST)
    if form.is_valid():
        intake.assigned_to = form.cleaned_data["assigned_to"]
        intake.save(update_fields=["assigned_to", "updated_at"])
        messages.success(request, "Хариуцагч шинэчлэгдлээ.")
    return redirect("dashboard:request_detail", code=code)


# ---------- Pickup ----------


@staff_required
def pickup_list(request):
    pickups = Pickup.objects.select_related("intake_request", "assigned_staff").order_by(
        "-pickup_date"
    )
    return render(request, "dashboard/pickup_list.html", {"pickups": pickups})


@staff_required
def pickup_detail(request, pk):
    pickup = get_object_or_404(Pickup, pk=pk)
    if request.method == "POST":
        form = PickupForm(request.POST, instance=pickup)
        if form.is_valid():
            form.save()
            messages.success(request, "Хадгалагдлаа.")
            return redirect("dashboard:pickup_detail", pk=pk)
    else:
        form = PickupForm(instance=pickup)
    return render(request, "dashboard/pickup_detail.html", {"pickup": pickup, "form": form})


@staff_required
def schedule_pickup(request, code):
    intake = get_object_or_404(IntakeRequest, request_code=code)
    if request.method == "POST":
        form = PickupForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            pickup, _ = Pickup.objects.update_or_create(
                intake_request=intake,
                defaults=dict(data),
            )
            notify_pickup_scheduled(pickup)
            messages.success(request, "Pickup товлогдлоо.")
            return redirect("dashboard:request_detail", code=code)
    else:
        form = PickupForm(initial={"pickup_address": intake.address_line})
    return render(
        request,
        "dashboard/schedule_pickup.html",
        {"form": form, "request_obj": intake},
    )


# ---------- Reports ----------


@staff_required
def reports(request):
    branches = Branch.objects.filter(is_active=True)
    statuses = IntakeRequest.Status.choices
    start, end = _period_range("this_month", timezone.localdate())
    return render(
        request,
        "dashboard/reports.html",
        {
            "branches": branches,
            "statuses": statuses,
            "date_from": start.isoformat(),
            "date_to": end.isoformat(),
            "period_choices": PERIOD_CHOICES,
        },
    )


@staff_required
def export(request):
    fmt = request.GET.get("format", "xlsx")
    qs = _build_export_queryset(request.GET)
    meta = _report_meta(request.GET, qs)

    if fmt == "csv":
        return _csv_export(qs)
    if fmt == "pdf":
        return _pdf_export(qs, meta)
    if fmt == "png":
        return _png_export(qs, meta)
    return _xlsx_export(qs)


def _report_meta(get_params, qs):
    """Human-readable title/subtitle for PDF & image reports (Mongolian)."""
    parts = []
    date_from = get_params.get("date_from")
    date_to = get_params.get("date_to")
    if date_from or date_to:
        parts.append(f"Хугацаа: {date_from or '…'} — {date_to or '…'}")
    if branch_id := get_params.get("branch"):
        from apps.branches.models import Branch

        branch = Branch.objects.filter(id=branch_id).first()
        if branch:
            parts.append(f"Салбар: {branch.name}")
    if statuses := [s for s in get_params.getlist("status") if s]:
        labels = dict(IntakeRequest.Status.choices)
        parts.append("Төлөв: " + ", ".join(labels.get(s, s) for s in statuses))
    subtitle = "  ·  ".join(parts) if parts else "Бүх захиалга"
    return {
        "title": "UBPM — Захиалгуудын тайлан",
        "subtitle": subtitle,
        "generated": timezone.localtime().strftime("%Y-%m-%d %H:%M"),
        "count": qs.count(),
    }


def _build_export_queryset(get_params):
    qs = IntakeRequest.objects.select_related("preferred_branch", "assigned_to").prefetch_related(
        "items", "pickup"
    )
    if branch := get_params.get("branch"):
        qs = qs.filter(preferred_branch_id=branch)
    if statuses := [s for s in get_params.getlist("status") if s]:
        qs = qs.filter(status__in=statuses)
    if date_from := get_params.get("date_from"):
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to := get_params.get("date_to"):
        qs = qs.filter(created_at__date__lte=date_to)
    return qs.order_by("-created_at")


def _row_for(req):
    actual = ""
    pickup = getattr(req, "pickup", None)
    if pickup and pickup.actual_buy_price:
        actual = float(pickup.actual_buy_price)
    return [
        req.request_code,
        req.created_at.strftime("%Y-%m-%d %H:%M"),
        req.contact_name,
        req.contact_phone,
        req.contact_email or "",
        req.company_name or "",
        req.preferred_branch.name if req.preferred_branch else "",
        req.get_status_display(),
        req.assigned_to.full_name if req.assigned_to else "",
        sum(i.quantity for i in req.items.all()),
        float(req.expected_price) if req.expected_price else "",
        actual,
    ]


def _xlsx_export(qs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Хүсэлтүүд"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="16A34A")
    for col, name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row, req in enumerate(qs, start=2):
        for col, val in enumerate(_row_for(req), 1):
            ws.cell(row=row, column=col, value=val)
    for col in range(1, len(EXCEL_COLUMNS) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"ubpm-requests-{timezone.localdate().isoformat()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _csv_export(qs):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    filename = f"ubpm-requests-{timezone.localdate().isoformat()}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write("\ufeff")  # BOM for Excel
    writer = csv.writer(response)
    writer.writerow(EXCEL_COLUMNS)
    for req in qs:
        writer.writerow(_row_for(req))
    return response


# ---------- PDF export (Mongolian via bundled DejaVu font) ----------

_PDF_FONTS_READY = False


def _ensure_pdf_fonts():
    global _PDF_FONTS_READY
    if _PDF_FONTS_READY:
        return
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    pdfmetrics.registerFont(TTFont("UBPM", str(settings.REPORT_FONT_REGULAR)))
    pdfmetrics.registerFont(TTFont("UBPM-Bold", str(settings.REPORT_FONT_BOLD)))
    pdfmetrics.registerFontFamily("UBPM", normal="UBPM", bold="UBPM-Bold")
    _PDF_FONTS_READY = True


def _pdf_export(qs, meta):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    _ensure_pdf_fonts()

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=meta["title"],
    )

    title_style = ParagraphStyle("t", fontName="UBPM-Bold", fontSize=16, leading=20)
    sub_style = ParagraphStyle("s", fontName="UBPM", fontSize=9, leading=13, textColor=colors.HexColor("#475569"))
    cell_style = ParagraphStyle("c", fontName="UBPM", fontSize=7, leading=9)
    head_style = ParagraphStyle("h", fontName="UBPM-Bold", fontSize=7.5, leading=9, textColor=colors.white)

    elements = [
        Paragraph(meta["title"], title_style),
        Paragraph(meta["subtitle"], sub_style),
        Paragraph(f"\u04ae\u04af\u0441\u0433\u044d\u0441\u044d\u043d: {meta['generated']}  \u00b7  \u041d\u0438\u0439\u0442: {meta['count']} \u0437\u0430\u0445\u0438\u0430\u043b\u0433\u0430", sub_style),
        Spacer(1, 6 * mm),
    ]

    header = [Paragraph(str(c), head_style) for c in EXCEL_COLUMNS]
    data = [header]
    for req in qs:
        data.append([Paragraph(_cell_text(v), cell_style) for v in _row_for(req)])

    # Column widths (mm) summing to ~273mm usable landscape A4 width.
    widths = [16, 24, 26, 20, 34, 26, 30, 26, 24, 14, 18, 18]
    table = Table(data, colWidths=[w * mm for w in widths], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16A34A")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buf.seek(0)

    response = HttpResponse(buf.getvalue(), content_type="application/pdf")
    filename = f"ubpm-requests-{timezone.localdate().isoformat()}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------- PNG image export (Mongolian via bundled DejaVu font) ----------

# (column label, pixel width)
_IMG_COLS = [
    ("\u041a\u043e\u0434", 90),
    ("\u041e\u0433\u043d\u043e\u043e", 130),
    ("\u0425\u044d\u0440\u044d\u0433\u043b\u044d\u0433\u0447", 140),
    ("\u0423\u0442\u0430\u0441", 95),
    ("Email", 180),
    ("\u041a\u043e\u043c\u043f\u0430\u043d\u0438", 130),
    ("\u0421\u0430\u043b\u0431\u0430\u0440", 170),
    ("\u0421\u0442\u0430\u0442\u0443\u0441", 150),
    ("\u0425\u0430\u0440\u0438\u0443\u0446\u0430\u0433\u0447", 130),
    ("\u0422\u04e9\u0445. \u0442\u043e\u043e", 80),
    ("\u0425\u04af\u0441\u0441\u044d\u043d \u04af\u043d\u044d", 100),
    ("\u0411\u043e\u0434\u0438\u0442 \u04af\u043d\u044d", 100),
]


def _fit_text(draw, text, font, max_width):
    """Truncate `text` with an ellipsis so it fits within `max_width` pixels."""
    if draw.textlength(text, font=font) <= max_width:
        return text
    ell = "\u2026"
    while text and draw.textlength(text + ell, font=font) > max_width:
        text = text[:-1]
    return text + ell


def _png_export(qs, meta):
    from PIL import Image, ImageDraw, ImageFont

    pad = 14
    row_h = 30
    header_block = 96
    col_widths = [w for _, w in _IMG_COLS]
    table_w = sum(col_widths)
    width = table_w + pad * 2

    rows = [_row_for(req) for req in qs]
    height = header_block + (len(rows) + 1) * row_h + pad

    font_title = ImageFont.truetype(str(settings.REPORT_FONT_BOLD), 26)
    font_sub = ImageFont.truetype(str(settings.REPORT_FONT_REGULAR), 13)
    font_head = ImageFont.truetype(str(settings.REPORT_FONT_BOLD), 13)
    font_cell = ImageFont.truetype(str(settings.REPORT_FONT_REGULAR), 12)

    img = Image.new("RGB", (width, height), "#ffffff")
    draw = ImageDraw.Draw(img)

    # Header text
    draw.text((pad, 12), meta["title"], font=font_title, fill="#14532d")
    draw.text((pad, 48), meta["subtitle"], font=font_sub, fill="#475569")
    draw.text(
        (pad, 68),
        f"\u04ae\u04af\u0441\u0433\u044d\u0441\u044d\u043d: {meta['generated']}   \u00b7   \u041d\u0438\u0439\u0442: {meta['count']} \u0437\u0430\u0445\u0438\u0430\u043b\u0433\u0430",
        font=font_sub,
        fill="#475569",
    )

    # Table header row
    y = header_block
    x = pad
    draw.rectangle([pad, y, pad + table_w, y + row_h], fill="#16A34A")
    for label, cw in _IMG_COLS:
        draw.text((x + 6, y + 7), _fit_text(draw, label, font_head, cw - 12), font=font_head, fill="#ffffff")
        x += cw

    # Data rows
    for i, row in enumerate(rows):
        y = header_block + (i + 1) * row_h
        if i % 2:
            draw.rectangle([pad, y, pad + table_w, y + row_h], fill="#f1f5f9")
        x = pad
        for value, cw in zip(row, col_widths, strict=False):
            draw.text(
                (x + 6, y + 8),
                _fit_text(draw, _cell_text(value), font_cell, cw - 12),
                font=font_cell,
                fill="#1f2937",
            )
            x += cw

    # Grid lines
    bottom = header_block + (len(rows) + 1) * row_h
    x = pad
    for cw in col_widths:
        draw.line([(x, header_block), (x, bottom)], fill="#cbd5e1", width=1)
        x += cw
    draw.line([(x, header_block), (x, bottom)], fill="#cbd5e1", width=1)
    for r in range(len(rows) + 2):
        yy = header_block + r * row_h
        draw.line([(pad, yy), (pad + table_w, yy)], fill="#cbd5e1", width=1)

    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="image/png")
    filename = f"ubpm-requests-{timezone.localdate().isoformat()}.png"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _cell_text(value):
    """Normalise a row value to a display string."""
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        return f"{value:,.0f}"
    return str(value)


# ---------------------------------------------------------------------------
# Email оношилгоо (зөвхөн ADMIN)
# ---------------------------------------------------------------------------
def _smtp_probe():
    """SMTP сервер рүү холбогдож нэвтэрнэ. Захиа ИЛГЭЭХГҮЙ.

    Буцаах: (ok, гарчиг, тайлбар). Тайлбар нь юуг засахыг шууд хэлнэ —
    Railway дээр CLI ажиллуулахгүйгээр шалтгааныг олох боломж.
    """
    import smtplib

    from django.core.mail import get_connection

    backend = getattr(settings, "EMAIL_BACKEND", "")
    if "console" in backend or "dummy" in backend:
        return (
            False,
            "EMAIL_BACKEND нь SMTP биш байна",
            "Захиа хэнд ч хүрэхгүй. RESEND_API_KEY ч, EMAIL_HOST_USER/"
            "EMAIL_HOST_PASSWORD ч тохируулаагүй үед prod автоматаар console руу "
            "унадаг. Railway → Variables дээр эдгээрийг нэмээд дахин deploy хийнэ үү.",
        )

    if "Resend" in backend:
        # HTTP API — SMTP холболт байхгүй тул түлхүүр тавигдсан эсэхийг л шалгана.
        if getattr(settings, "RESEND_API_KEY", ""):
            return (
                True,
                "Resend (HTTP API) ашиглаж байна",
                "SMTP порт шаардлагагүй тул Railway-гийн хоригт өртөхгүй.\n"
                f"From хаяг: {settings.DEFAULT_FROM_EMAIL}\n\n"
                "Захиа очихгүй бол доорх бүртгэлээс Resend-ийн хариуг хараарай — "
                "домайн баталгаажаагүй үед 403 буцаадаг.",
            )
        return (
            False,
            "RESEND_API_KEY хоосон байна",
            "Railway → Variables дээр RESEND_API_KEY-гээ нэмнэ үү.",
        )

    connection = None
    try:
        connection = get_connection()
        connection.open()
        return (True, "Холбогдож, нэвтэрлээ", "SMTP тохиргоо хэвийн ажиллаж байна.")
    except smtplib.SMTPAuthenticationError as exc:
        return (
            False,
            "Gmail нэвтрэлт амжилтгүй",
            f"{exc!r}\n\nEMAIL_HOST_PASSWORD нь энгийн нууц үг биш, Google Account → "
            "Security → 2-Step Verification → App passwords дээрээс үүсгэсэн "
            "16 тэмдэгт App Password байх ёстой.",
        )
    except smtplib.SMTPException as exc:
        return (False, "SMTP серверийн алдаа", repr(exc))
    except (TimeoutError, OSError) as exc:
        return (
            False,
            "SMTP сервер рүү холбогдож чадсангүй",
            f"{exc!r}\n\nИхэвчлэн hosting тал нь гадагш SMTP портыг хаасан үед ингэдэг. "
            "EMAIL_PORT=465 + EMAIL_USE_SSL=True туршиж үзэх, эсвэл HTTP API-тай "
            "email үйлчилгээ (Resend, SendGrid) руу шилжих шаардлагатай.",
        )
    except Exception as exc:  # noqa: BLE001
        return (False, "SMTP холболт амжилтгүй", repr(exc))
    finally:
        if connection is not None:
            # Холболт хаагдахгүй байх нь оношилгооны хариуг өөрчлөхгүй.
            with contextlib.suppress(Exception):
                connection.close()


@role_required(User.Role.ADMIN)
def email_status(request):
    """Email тохиргоог браузераас шалгах хуудас (Railway CLI-гүйгээр)."""
    password = getattr(settings, "EMAIL_HOST_PASSWORD", "")
    config = [
        ("EMAIL_BACKEND", getattr(settings, "EMAIL_BACKEND", "")),
        ("EMAIL_HOST", getattr(settings, "EMAIL_HOST", "")),
        ("EMAIL_PORT", getattr(settings, "EMAIL_PORT", "")),
        ("EMAIL_USE_TLS", getattr(settings, "EMAIL_USE_TLS", False)),
        ("EMAIL_USE_SSL", getattr(settings, "EMAIL_USE_SSL", False)),
        ("EMAIL_HOST_USER", getattr(settings, "EMAIL_HOST_USER", "") or "— хоосон —"),
        # Нууц үгийг хэвлэхгүй — тавигдсан эсэх, урт нь л хангалттай мэдээлэл.
        (
            "EMAIL_HOST_PASSWORD",
            f"тавигдсан ({len(password)} тэмдэгт)" if password else "— хоосон —",
        ),
        ("DEFAULT_FROM_EMAIL", settings.DEFAULT_FROM_EMAIL),
        ("EMAIL_TIMEOUT", getattr(settings, "EMAIL_TIMEOUT", None)),
        ("EMAIL_ASYNC", getattr(settings, "EMAIL_ASYNC", True)),
        ("SITE_URL", getattr(settings, "SITE_URL", "") or "— хоосон —"),
    ]

    probe = None
    if request.method == "POST":
        probe = _smtp_probe()

    return render(
        request,
        "dashboard/email_status.html",
        {
            "config": config,
            "probe": probe,
            "recent": EmailLog.objects.all()[:20],
            "failed_count": EmailLog.objects.filter(success=False).count(),
        },
    )
